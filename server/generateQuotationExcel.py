#!/usr/bin/env python3
"""
Generate professional quotation Excel files with openpyxl
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_quotation_excel(quotation_data: dict, items_data: list) -> bytes:
    """
    Generate a professional quotation Excel file
    
    Args:
        quotation_data: Dictionary with quotation info (customerName, projectName, etc.)
        items_data: List of product items
    
    Returns:
        Excel file as bytes
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"
    
    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # Define styles
    title_font = Font(name='黑体', size=20, bold=True, color='1F1F1F')
    header_font = Font(name='黑体', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    info_font = Font(name='黑体', size=11, bold=True)
    info_alignment = Alignment(horizontal='left', vertical='center')
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_alignment_right = Alignment(horizontal='right', vertical='center')
    
    total_font = Font(name='黑体', size=11, bold=True, color='1F1F1F')
    total_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    total_border = Border(
        top=Side(style='medium', color='7C3AED'),
        bottom=Side(style='medium', color='7C3AED')
    )
    
    light_border = Border(
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0')
    )
    
    # Row 1-2: Empty for logo space
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # Row 3: Title
    ws.row_dimensions[3].height = 30
    ws['A3'] = 'DAN 产品报价单'
    ws['A3'].font = title_font
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Row 4: Empty
    ws.row_dimensions[4].height = 10
    
    # Row 5: Customer Info Header
    ws.row_dimensions[5].height = 20
    ws['A5'] = '客户信息'
    ws['A5'].font = info_font
    ws['A5'].alignment = info_alignment
    
    # Row 6: Customer Details - Line 1
    ws.row_dimensions[6].height = 20
    ws['A6'] = '报价编号'
    ws['B6'] = quotation_data.get('quotationNo', '')
    ws['D6'] = '客户名称'
    ws['E6'] = quotation_data.get('customerName', '')
    for col in ['A', 'B', 'D', 'E']:
        ws[f'{col}6'].font = data_font
        ws[f'{col}6'].alignment = data_alignment
    
    # Row 7: Customer Details - Line 2
    ws.row_dimensions[7].height = 20
    ws['A7'] = '项目名称'
    ws['B7'] = quotation_data.get('projectName', '')
    ws['D7'] = '联系人'
    ws['E7'] = quotation_data.get('customerContact', '')
    for col in ['A', 'B', 'D', 'E']:
        ws[f'{col}7'].font = data_font
        ws[f'{col}7'].alignment = data_alignment
    
    # Row 8: Customer Details - Line 3
    ws.row_dimensions[8].height = 20
    ws['A8'] = '电话'
    ws['B8'] = quotation_data.get('customerPhone', '')
    ws['D8'] = '邮箱'
    ws['E8'] = quotation_data.get('customerEmail', '')
    for col in ['A', 'B', 'D', 'E']:
        ws[f'{col}8'].font = data_font
        ws[f'{col}8'].alignment = data_alignment
    
    # Row 9: Customer Details - Line 4
    ws.row_dimensions[9].height = 20
    created_date = ''
    if quotation_data.get('createdAt'):
        created_date = datetime.fromisoformat(quotation_data['createdAt'].replace('Z', '+00:00')).strftime('%Y-%m-%d')
    valid_until = ''
    if quotation_data.get('validUntil'):
        valid_until = datetime.fromisoformat(quotation_data['validUntil'].replace('Z', '+00:00')).strftime('%Y-%m-%d')
    
    ws['A9'] = '创建日期'
    ws['B9'] = created_date
    ws['D9'] = '有效期'
    ws['E9'] = valid_until
    for col in ['A', 'B', 'D', 'E']:
        ws[f'{col}9'].font = data_font
        ws[f'{col}9'].alignment = data_alignment
    
    # Row 10: Empty
    ws.row_dimensions[10].height = 10
    
    # Row 11: Product Header
    ws.row_dimensions[11].height = 25
    headers = ['序号', '产品型号', '产品说明', '单价(¥)', '数量', '折扣率(%)', '小计(¥)', '媒体价(¥)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = light_border
    
    # Product data rows
    total = 0
    for item_idx, item in enumerate(items_data):
        row_num = 12 + item_idx
        ws.row_dimensions[row_num].height = 25
        
        list_price = float(item.get('listPrice', 0)) if item.get('listPrice') else 0
        discount = float(item.get('discountRate', 0)) if item.get('discountRate') else 0
        unit_price = list_price * (1 - discount / 100)
        qty = int(item.get('quantity', 1)) if item.get('quantity') else 1
        subtotal = unit_price * qty
        total += subtotal
        
        # Row data
        row_data = [
            item_idx + 1,
            item.get('productModel', ''),
            item.get('productDesc', ''),
            f'{unit_price:.2f}',
            qty,
            f'{discount:.0f}',
            f'{subtotal:.2f}',
            f'{list_price:.2f}' if list_price > 0 else '',
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.border = light_border
            
            # Alignment
            if col_idx >= 4:  # Numeric columns
                cell.alignment = data_alignment_right
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=col_idx == 3)
            
            # Alternating row background
            if item_idx % 2 == 0:
                cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
    
    # Total row
    total_row = 12 + len(items_data) + 1
    ws.row_dimensions[total_row].height = 25
    
    ws[f'A{total_row}'] = '合计'
    ws[f'A{total_row}'].font = total_font
    ws[f'A{total_row}'].fill = total_fill
    ws[f'A{total_row}'].border = total_border
    ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'G{total_row}'] = f'{total:.2f}'
    ws[f'G{total_row}'].font = total_font
    ws[f'G{total_row}'].fill = total_fill
    ws[f'G{total_row}'].border = total_border
    ws[f'G{total_row}'].alignment = data_alignment_right
    
    # Freeze panes
    ws.freeze_panes = 'A12'
    
    # Save to bytes
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

if __name__ == '__main__':
    # Read from stdin
    input_data = json.loads(sys.stdin.read())
    quotation = input_data.get('quotation', {})
    items = input_data.get('items', [])
    
    # Generate Excel
    excel_bytes = generate_quotation_excel(quotation, items)
    
    # Write to stdout as base64
    import base64
    print(base64.b64encode(excel_bytes).decode('utf-8'))
